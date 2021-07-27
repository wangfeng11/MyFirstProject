#!/usr/bin/env python
# -*- conding:utf-8 -*-
# author:wangfeng
from openpyxl import load_workbook

wb = load_workbook(filename='empty_book.xlsx')
sheet_ranges = wb['range names']
sheet_ranges_mySheet = wb['my_Sheet']
print(sheet_ranges['D18'].value)
for i in range(1, 30):
    print(sheet_ranges_mySheet.cell(column=1, row=i).value)
